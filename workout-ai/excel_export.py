import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from schemas import WorkoutPlan, UserInput


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _apply_border(ws, min_row, max_row, min_col, max_col):
    border = _thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


# ---------------------------------------------------------------------------
# Sheet 1 — Schedules
# ---------------------------------------------------------------------------

def _build_schedules_sheet(wb: Workbook, plan: WorkoutPlan, user: UserInput):
    ws = wb.active
    ws.title = "Schedules"
    ws.sheet_properties.tabColor = "2E75B6"

    NUM_COLS = 8  # Ngày | Nhóm cơ | Bài tập | Sets | Reps | Nghỉ | Hướng dẫn | Lưu ý

    # --- Row 1: Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1, value="LỊCH TẬP TUẦN")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _fill("2E75B6")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Row 2: Subtitle ---
    subtitle = (
        f"Mục tiêu: {user.goal_type.value} | "
        f"Calories: {plan.daily_calories} kcal | "
        f"Protein: {plan.protein_g}g"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.fill = _fill("BDD7EE")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # --- Row 3: Spacer ---
    ws.row_dimensions[3].height = 8

    # --- Row 4: Headers ---
    headers = ["Ngày", "Nhóm cơ", "Bài tập", "Sets", "Reps", "Nghỉ", "Hướng dẫn", "Lưu ý"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill("1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = "A5"

    # Column widths
    col_widths = [12, 18, 20, 8, 10, 12, 45, 35]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Build ordered day list: workout days interleaved with rest days
    rest_days = set(plan.rest_days)
    # Canonical week order in Vietnamese
    week_order = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
    workout_by_day = {d.day: d for d in plan.workout_plan}

    # Build the ordered sequence of all days (workout + rest)
    all_days = []
    for day in week_order:
        if day in workout_by_day or day in rest_days:
            all_days.append(day)
    # Add any workout days not in the canonical order
    for d in plan.workout_plan:
        if d.day not in all_days:
            all_days.append(d.day)

    # Alternate fill colors per workout day
    day_fills = ["DEEAF1", "FFFFFF"]
    fill_index = 0

    current_row = 5
    for day_name in all_days:
        if day_name in rest_days and day_name not in workout_by_day:
            # Rest day row
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=NUM_COLS,
            )
            cell = ws.cell(
                row=current_row, column=1,
                value="NGHỈ NGƠI — Phục hồi cơ thể",
            )
            cell.font = Font(italic=True)
            cell.fill = _fill("FFF2CC")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _apply_border(ws, current_row, current_row, 1, NUM_COLS)
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        elif day_name in workout_by_day:
            day_data = workout_by_day[day_name]
            day_fill_color = day_fills[fill_index % 2]
            fill_index += 1
            num_exercises = len(day_data.exercises)
            day_start_row = current_row

            for ex_idx, exercise in enumerate(day_data.exercises):
                row = current_row + ex_idx
                row_fill = _fill(day_fill_color)

                ws.cell(row=row, column=3, value=exercise.name).fill = row_fill
                ws.cell(row=row, column=4, value=exercise.sets).fill = row_fill
                ws.cell(row=row, column=5, value=exercise.reps).fill = row_fill
                ws.cell(row=row, column=6, value=exercise.rest).fill = row_fill

                inst_cell = ws.cell(row=row, column=7, value=exercise.instructions)
                inst_cell.fill = row_fill
                inst_cell.alignment = Alignment(wrap_text=True, vertical="top")

                notes_cell = ws.cell(row=row, column=8, value=exercise.notes)
                notes_cell.fill = row_fill
                notes_cell.alignment = Alignment(wrap_text=True, vertical="top")

                for col in [3, 4, 5, 6]:
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal="center", vertical="top"
                    )

                ws.row_dimensions[row].height = 60

            # Merge Ngày and Nhóm cơ vertically across all exercises
            end_row = current_row + num_exercises - 1

            if num_exercises > 1:
                ws.merge_cells(
                    start_row=day_start_row, start_column=1,
                    end_row=end_row, end_column=1,
                )
                ws.merge_cells(
                    start_row=day_start_row, start_column=2,
                    end_row=end_row, end_column=2,
                )

            day_cell = ws.cell(row=day_start_row, column=1, value=day_name)
            day_cell.fill = _fill(day_fill_color)
            day_cell.font = Font(bold=True)
            day_cell.alignment = Alignment(horizontal="center", vertical="top")

            focus_cell = ws.cell(row=day_start_row, column=2, value=day_data.focus)
            focus_cell.fill = _fill(day_fill_color)
            focus_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

            _apply_border(ws, day_start_row, end_row, 1, NUM_COLS)
            current_row += num_exercises


# ---------------------------------------------------------------------------
# Sheet 2 — Meals
# ---------------------------------------------------------------------------

def _build_meals_sheet(wb: Workbook, plan: WorkoutPlan):
    ws = wb.create_sheet("Meals")
    ws.sheet_properties.tabColor = "548235"

    NUM_COLS = 5  # Ngày | Bữa ăn | Thực phẩm | Calories | Protein (g)

    # --- Row 1: Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1, value="THỰC ĐƠN TUẦN")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _fill("548235")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Row 2: Subtitle ---
    subtitle = (
        f"Calories/ngày: {plan.daily_calories} kcal | "
        f"Protein: {plan.protein_g}g | "
        f"Carbs: {plan.carbs_g}g | "
        f"Fat: {plan.fat_g}g"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.fill = _fill("E2EFDA")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # --- Row 3: Spacer ---
    ws.row_dimensions[3].height = 8

    # --- Row 4: Headers ---
    headers = ["Ngày", "Bữa ăn", "Thực phẩm", "Calories", "Protein (g)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill("375623")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = "A5"

    # Column widths
    col_widths = [22, 20, 40, 14, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Meal type background colors
    meal_colors = {
        "sáng": "FFF2CC",
        "trưa": "DEEAF1",
        "tối": "FCE4D6",
        "snack": "E2EFDA",
        "phụ": "E2EFDA",
    }

    def meal_color(meal_name: str) -> str:
        name_lower = meal_name.lower()
        for keyword, color in meal_colors.items():
            if keyword in name_lower:
                return color
        return "FFFFFF"

    current_row = 5

    for day_data in plan.meal_plan:
        day_start_row = current_row
        num_meals = len(day_data.meals)
        total_cal = 0
        total_prot = 0

        for meal in day_data.meals:
            color = meal_color(meal.meal)
            row_fill = _fill(color)

            meal_cell = ws.cell(row=current_row, column=2, value=meal.meal)
            meal_cell.fill = row_fill
            meal_cell.alignment = Alignment(vertical="top", wrap_text=True)

            foods_text = "\n".join(meal.foods)
            foods_cell = ws.cell(row=current_row, column=3, value=foods_text)
            foods_cell.alignment = Alignment(wrap_text=True, vertical="top")

            cal_cell = ws.cell(row=current_row, column=4, value=meal.calories)
            cal_cell.alignment = Alignment(horizontal="center", vertical="top")

            prot_cell = ws.cell(row=current_row, column=5, value=meal.protein_g)
            prot_cell.alignment = Alignment(horizontal="center", vertical="top")

            ws.row_dimensions[current_row].height = 60
            total_cal += meal.calories
            total_prot += meal.protein_g
            current_row += 1

        # Subtotal row
        subtotal_fill = _fill("E2EFDA")
        ws.cell(row=current_row, column=2, value="TỔNG NGÀY").font = Font(bold=True)
        ws.cell(row=current_row, column=2).fill = subtotal_fill
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=current_row, column=3).fill = subtotal_fill

        cal_total = ws.cell(row=current_row, column=4, value=total_cal)
        cal_total.font = Font(bold=True)
        cal_total.fill = subtotal_fill
        cal_total.alignment = Alignment(horizontal="center", vertical="center")

        prot_total = ws.cell(row=current_row, column=5, value=total_prot)
        prot_total.font = Font(bold=True)
        prot_total.fill = subtotal_fill
        prot_total.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[current_row].height = 18
        subtotal_row = current_row
        current_row += 1

        # Merge Ngày column across all meal rows + subtotal
        day_end_row = subtotal_row
        if day_end_row > day_start_row:
            ws.merge_cells(
                start_row=day_start_row, start_column=1,
                end_row=day_end_row, end_column=1,
            )

        day_cell = ws.cell(row=day_start_row, column=1, value=day_data.day)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        _apply_border(ws, day_start_row, day_end_row, 1, NUM_COLS)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel(plan: WorkoutPlan, user: UserInput) -> io.BytesIO:
    wb = Workbook()
    _build_schedules_sheet(wb, plan, user)
    _build_meals_sheet(wb, plan)

    # Remove the default empty sheet openpyxl creates (already replaced by _build_schedules_sheet
    # which re-uses wb.active, so nothing to delete here)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
